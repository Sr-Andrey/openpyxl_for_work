import os
from openpyxl import load_workbook

current_directory = os.getcwd()
name_files1 = input('Введите имя первого файла: ')+'.xlsx'
wb = load_workbook(current_directory+'\\'+name_files1)
wb.active = wb['Макет отчета']
sheet = wb.active
pok_1 = {}
x, y = map(int, input("Введите строку начала и конца исходника: ",).split())
for i in range(x, y+1):
    pok_1[str(sheet.cell(i, 10).value)] = 0
    li = []
    for j in range(11, 13):
        li.append(str(sheet.cell(i, j).value))
    if sheet.cell(i, 13).value == None:
        li.append(0)
    else:
        li.append(sheet.cell(i, 13).value)

    pok_1[str(sheet.cell(i, 10).value)] = li

wb.close()

name_files1 = input('Введите имя второго файла: ')+'.xlsx'
wb = load_workbook(current_directory+'\\'+name_files1)
wb.active = wb['Макет отчета']
sheet = wb.active
pokk_2 = {}
x, y = map(int, input("Введите строку начала и конца исходника: ",).split())
for i in range(x, y+1):
    pokk_2[str(sheet.cell(i, 10).value)] = 0
    li = []
    for j in range(11, 13):
        li.append(str(sheet.cell(i, j).value))
    if sheet.cell(i, 13).value == None:
        li.append(0)
    else:
        li.append(sheet.cell(i, 13).value)
    pokk_2[str(sheet.cell(i, 10).value)] = li
