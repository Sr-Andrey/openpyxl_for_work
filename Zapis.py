import os
from openpyxl import load_workbook
from raschet_d import pok_1
from raschet_c import pokk_2
from raschet_d import name_files1
from openpyxl.styles import Color, PatternFill


def main():
        current_directory = os.getcwd()
        
        wb=load_workbook(current_directory+'\\'+name_files1)

        wb.create_sheet(title = 'ТЭР Д', index = 0)
        sheet=wb['ТЭР Д']
        sheet.cell(1,1).value = '№ пп'
        sheet.cell(1,2).value = 'Наименование бюджетного показателя'
        sheet.cell(1,3).value = 'Ед.изм.'
        sheet.cell(1,4).value = 'Факт'

        x=2
        for k,v in pok_1.items():
                sheet.cell(x,1).value=k
                sheet.cell(x,2).value=v[0]
                sheet.cell(x,3).value=v[1]
                sheet.cell(x,4).value=v[2]
                x+=1
        redFill = PatternFill(start_color='00aaff',
                        end_color='00aaff',
                        fill_type='solid')
        for i in range(31): 
                if i in [9,10,12,13,22,24,25,26,28,29,30]: 
                        sheet.cell(i,4).fill = redFill


        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['A'].width = 20

        wb.create_sheet(title = 'ТЭР Ц', index = 0)
        sheet=wb['ТЭР Ц']
        sheet.cell(1,1).value = '№ пп'
        sheet.cell(1,2).value = 'Наименование бюджетного показателя'
        sheet.cell(1,3).value = 'Ед.изм.'
        sheet.cell(1,4).value = 'Факт'

        x=2
        for k,v in pokk_2.items():
                sheet.cell(x,1).value=k
                sheet.cell(x,2).value=v[0]
                sheet.cell(x,3).value=v[1]
                sheet.cell(x,4).value=v[2]
                x+=1

        redFill = PatternFill(start_color='00aaff',
                        end_color='00aaff',
                        fill_type='solid')
        for i in range(65): 
                if i in [5,6,8,9,11,12,14,17,18,19,21,22,24,25,36,45,47,48,50,51,53,54,55,56,58,59,60,62,63,64]: 
                        sheet.cell(i,4).fill = redFill
                        
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['A'].width = 20

        wb.save('TER_ready.xlsx')
        print('Finish')

        


if __name__=="__main__": main()