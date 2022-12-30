import sys
import os
from openpyxl import load_workbook
from Sbor import pokk_2
from raschet_d import name_files1
current_directory = os.getcwd()

wb = load_workbook(current_directory+'\\'+name_files1)
x, y, l = map(str, input("Введите строку и столбец брутто и номер листа элетро книги \
Факт всего, через пробел, пример 110 aj 3: ",).split())
# x,y,l=110,'aj',3
x, l = int(x), int(l)
list_stol = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
             'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ',
             'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ', 'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ', 'FK', 'FL', 'FM', 'FN', 'FO',
             'FP', 'FQ', 'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ', 'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN', 'GO', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ', 'HA', 'HB', 'HC', 'HD', 'HE', 'HF', 'HG', 'HH', 'HI', 'HJ', 'HK', 'HL', 'HM', 'HN', 'HO', 'HP', 'HQ', 'HR', 'HS', 'HT', 'HU', 'HV', 'HW', 'HX', 'HY', 'HZ', 'IA', 'IB', 'IC', 'ID',
             'IE', 'IF', 'IG', 'IH', 'II', 'IJ', 'IK', 'IL', 'IM', 'IN', 'IO', 'IP', 'IQ', 'IR', 'IS', 'IT', 'IU', 'IV', 'IW', 'IX', 'IY', 'IZ', 'JA', 'JB', 'JC', 'JD', 'JE', 'JF', 'JG', 'JH', 'JI', 'JJ', 'JK', 'JL', 'JM', 'JN', 'JO', 'JP', 'JQ', 'JR', 'JS', 'JT', 'JU', 'JV', 'JW', 'JX', 'JY', 'JZ', 'KA', 'KB', 'KC', 'KD', 'KE', 'KF', 'KG', 'KH', 'KI', 'KJ', 'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ', 'KR', 'KS',
             'KT', 'KU', 'KV', 'KW', 'KX', 'KY', 'KZ', 'LA', 'LB', 'LC', 'LD', 'LE', 'LF', 'LG', 'LH', 'LI', 'LJ', 'LK', 'LL', 'LM', 'LN', 'LO', 'LP', 'LQ', 'LR', 'LS', 'LT', 'LU', 'LV', 'LW', 'LX', 'LY', 'LZ', 'MA', 'MB', 'MC', 'MD', 'ME', 'MF', 'MG', 'MH', 'MI', 'MJ', 'MK', 'ML', 'MM', 'MN', 'MO', 'MP', 'MQ', 'MR', 'MS', 'MT', 'MU', 'MV', 'MW', 'MX', 'MY', 'MZ', 'NA', 'NB', 'NC', 'ND', 'NE', 'NF', 'NG', 'NH',
             'NI', 'NJ', 'NK', 'NL', 'NM', 'NN', 'NO', 'NP', 'NQ', 'NR', 'NS', 'NT', 'NU', 'NV', 'NW', 'NX', 'NY', 'NZ', 'OA', 'OB', 'OC', 'OD', 'OE', 'OF', 'OG', 'OH', 'OI', 'OJ', 'OK', 'OL', 'OM', 'ON', 'OO', 'OP', 'OQ', 'OR', 'OS', 'OT', 'OU', 'OV', 'OW', 'OX', 'OY', 'OZ', 'PA', 'PB', 'PC', 'PD', 'PE', 'PF', 'PG', 'PH', 'PI', 'PJ', 'PK', 'PL', 'PM', 'PN', 'PO', 'PP', 'PQ', 'PR', 'PS', 'PT', 'PU', 'PV', 'PW',
             'PX', 'PY', 'PZ', 'QA', 'QB', 'QC', 'QD', 'QE', 'QF', 'QG', 'QH', 'QI', 'QJ', 'QK', 'QL', 'QM', 'QN', 'QO', 'QP', 'QQ', 'QR', 'QS', 'QT', 'QU', 'QV', 'QW', 'QX', 'QY', 'QZ', 'RA', 'RB', 'RC', 'RD', 'RE', 'RF', 'RG', 'RH', 'RI', 'RJ', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RQ', 'RR', 'RS', 'RT', 'RU', 'RV', 'RW', 'RX', 'RY', 'RZ', 'SA', 'SB', 'SC', 'SD', 'SE', 'SF', 'SG', 'SH', 'SI', 'SJ', 'SK', 'SL',
             'SM', 'SN', 'SO', 'SP', 'SQ', 'SR', 'SS', 'ST', 'SU', 'SV', 'SW', 'SX', 'SY', 'SZ', 'TA', 'TB', 'TC', 'TD', 'TE', 'TF', 'TG', 'TH', 'TI', 'TJ', 'TK', 'TL', 'TM', 'TN', 'TO', 'TP', 'TQ', 'TR', 'TS', 'TT', 'TU', 'TV', 'TW', 'TX', 'TY', 'TZ', 'UA', 'UB', 'UC', 'UD', 'UE', 'UF', 'UG', 'UH', 'UI', 'UJ', 'UK', 'UL', 'UM', 'UN', 'UO', 'UP', 'UQ', 'UR', 'US', 'UT', 'UU', 'UV', 'UW', 'UX', 'UY', 'UZ', 'VA',
             'VB', 'VC', 'VD', 'VE', 'VF', 'VG', 'VH', 'VI', 'VJ', 'VK', 'VL', 'VM', 'VN', 'VO', 'VP', 'VQ', 'VR', 'VS', 'VT', 'VU', 'VV', 'VW', 'VX', 'VY', 'VZ', 'WA', 'WB', 'WC', 'WD', 'WE', 'WF', 'WG', 'WH', 'WI', 'WJ', 'WK', 'WL', 'WM', 'WN', 'WO', 'WP', 'WQ', 'WR', 'WS', 'WT', 'WU', 'WV', 'WW', 'WX', 'WY', 'WZ', 'XA', 'XB', 'XC', 'XD', 'XE', 'XF', 'XG', 'XH', 'XI', 'XJ', 'XK', 'XL', 'XM', 'XN', 'XO', 'XP',
             'XQ', 'XR', 'XS', 'XT', 'XU', 'XV', 'XW', 'XX', 'XY', 'XZ', 'YA', 'YB', 'YC', 'YD', 'YE', 'YF', 'YG', 'YH', 'YI', 'YJ', 'YK', 'YL', 'YM', 'YN', 'YO', 'YP', 'YQ', 'YR', 'YS', 'YT', 'YU', 'YV', 'YW', 'YX', 'YY', 'YZ', 'ZA', 'ZB', 'ZC', 'ZD', 'ZE', 'ZF', 'ZG', 'ZH', 'ZI', 'ZJ', 'ZK', 'ZL', 'ZM', 'ZN', 'ZO', 'ZP', 'ZQ', 'ZR', 'ZS', 'ZT', 'ZU', 'ZV', 'ZW', 'ZX', 'ZY', 'ZZ']

y = list_stol.index(y.upper())
wb.active = l-1
sheet = wb.active
# brutto vsego
pokk_2['2.6.01.2'][2] = sheet.cell(x, y).value
pokk_2['2.6.03.3'][2] = sheet.cell(x, y+1).value
pokk_2['2.6.03.5'][2] = (pokk_2['2.6.01.2'][2]*pokk_2['2.6.03.3'][2])/10000
# brutto xoz
pokk_2['2.6.01.2.в'][2] = sheet.cell(x, y+9).value
pokk_2['2.6.03.3.в'][2] = sheet.cell(x, y+10).value
pokk_2['2.6.03.5.в'][2] = (pokk_2['2.6.01.2.в'][2]
                           * pokk_2['2.6.03.3.в'][2])/10000

# brutto dalnee
pasdal = float(input('ВВедите значение электротяга, пример - 3837.8 '))/100
# pasdal=3837.8/100
pokk_2['2.6.01.2.б.1'][2] = pasdal
pokk_2['2.6.03.3.б.1'][2] = sheet.cell(x, y+18).value
pokk_2['2.6.03.5.б.1'][2] = (
    pokk_2['2.6.01.2.б.1'][2]*pokk_2['2.6.03.3.б.1'][2])/10000
# brutto prigod
pokk_2['2.6.01.2.б.2'][2] = sheet.cell(x, y+17).value - pasdal
pokk_2['2.6.03.3.б.2'][2] = sheet.cell(x, y+18).value
pokk_2['2.6.03.5.б.2'][2] = (
    pokk_2['2.6.01.2.б.2'][2]*pokk_2['2.6.03.3.б.2'][2])/10000

pokk_2['2.6.01.2.б'][2] = pokk_2['2.6.01.2.б.1'][2]+pokk_2['2.6.01.2.б.2'][2]
pokk_2['2.6.01.2.в.1'][2] = pokk_2['2.6.01.2.в'][2]
pokk_2['2.6.03.3.в.1'][2] = pokk_2['2.6.03.3.в'][2]
pokk_2['2.6.03.5.в.1'][2] = pokk_2['2.6.03.5.в'][2]
pokk_2['2.6.03.5.б'][2] = pokk_2['2.6.03.5.б.1'][2]+pokk_2['2.6.03.5.б.2'][2]

prov_brutto = pokk_2['2.6.01.2'][2]-pokk_2['2.6.01.2.б.1'][2] - \
    pokk_2['2.6.01.2.б.2'][2]-pokk_2['2.6.01.2.в'][2]
prov_potreblenie = pokk_2['2.6.03.5'][2]-pokk_2['2.6.03.5.б.1'][2] - \
    pokk_2['2.6.03.5.б.2'][2]-pokk_2['2.6.03.5.в'][2]

if prov_brutto > 0.1 or prov_potreblenie > 0.01:
    print("Ошибка сумм по видам  Электро ", prov_brutto, prov_potreblenie)
    sys.exit()
wb.close()
x, y, l = map(str, input("Введите строку и столбец брутто и номер листа тепло книги \
Факт всего, через пробел, пример 90 с 2: ",).split())
# x,y,l=90,'c',2
x, l = int(x), int(l)
y = list_stol.index(y.upper())

wb.active = l-1
sheet = wb.active

pasdal = float(
    input('ВВедите значени дальнее теплотяга, пример - 293.454420000001 '))/100
# pasdal=293.454420000001/100

proverka1_pot = (sheet.cell(x, y+11).value*sheet.cell(x, y+12).value)/10
proverka2_pot = (sheet.cell(x, y+13).value*sheet.cell(x, y+14).value)/10
proverka3_pot = proverka1_pot+proverka2_pot

proverka1_bryt = sheet.cell(x, y+11).value+sheet.cell(x, y+13).value
proverka_udel_ras = proverka3_pot/proverka1_bryt*10

# brutto vsego
pokk_2['2.6.01.1'][2] = sheet.cell(x, y).value
pokk_2['2.6.03.1'][2] = sheet.cell(x, y+1).value
pokk_2['2.6.03.4'][2] = (pokk_2['2.6.01.1'][2]*pokk_2['2.6.03.1'][2])/10
# brutto hoz
pokk_2['2.6.01.1.в'][2] = sheet.cell(x, y+6).value
pokk_2['2.6.03.1.в'][2] = sheet.cell(x, y+7).value
pokk_2['2.6.03.4.в'][2] = (pokk_2['2.6.01.1.в'][2]*pokk_2['2.6.03.1.в'][2])/10
# brutto manev
pokk_2['2.6.01.1.0'][2] = sheet.cell(x, y+8).value
pokk_2['2.6.03.1.г'][2] = sheet.cell(x, y+10).value
pokk_2['2.6.03.4.г'][2] = (pokk_2['2.6.01.1.0'][2]*pokk_2['2.6.03.1.г'][2])*10
# brutto prig
pokk_2['2.6.01.1.б.2'][2] = proverka1_bryt-pasdal
pokk_2['2.6.03.1.б.2'][2] = proverka_udel_ras
pokk_2['2.6.03.4.б.2'][2] = (
    pokk_2['2.6.01.1.б.2'][2]*pokk_2['2.6.03.1.б.2'][2])/10
# brutto dal
pokk_2['2.6.01.1.б.1'][2] = pasdal
pokk_2['2.6.03.1.б.1'][2] = proverka_udel_ras
pokk_2['2.6.03.4.б.1'][2] = (
    pokk_2['2.6.01.1.б.1'][2]*pokk_2['2.6.03.1.б.1'][2])/10

pokk_2['2.6.01.1.б'][2] = pokk_2['2.6.01.1.б.1'][2]+pokk_2['2.6.01.1.б.2'][2]
pokk_2['2.6.01.1.б.1.2'][2] = pokk_2['2.6.01.1.б.1'][2]
pokk_2['2.6.01.1.в.1'][2] = pokk_2['2.6.01.1.в'][2]
pokk_2['2.6.01.1.г'][2] = pokk_2['2.6.01.1'][2]-pokk_2['2.6.01.1.б.1'][2] - \
    pokk_2['2.6.01.1.б.2'][2]-pokk_2['2.6.01.1.в'][2]
pokk_2['2.6.03.1.б.1.2'][2] = pokk_2['2.6.03.1.б.1'][2]
pokk_2['2.6.03.1.в.1'][2] = pokk_2['2.6.03.1.в'][2]
pokk_2['2.6.03.4.б'][2] = pokk_2['2.6.03.4.б.1'][2]+pokk_2['2.6.03.4.б.2'][2]
pokk_2['2.6.03.4.б.1.2'][2] = pokk_2['2.6.03.4.б.1'][2]
pokk_2['2.6.03.4.в.1'][2] = pokk_2['2.6.03.4.в'][2]

prov_potreblenie = pokk_2['2.6.03.4'][2]-pokk_2['2.6.03.4.б.1'][2] - \
    pokk_2['2.6.03.4.б.2'][2]-pokk_2['2.6.03.4.в'][2]-pokk_2['2.6.03.4.г'][2]
if prov_potreblenie > 0.01:
    print("Ошибка сумм по видам движения теплотяга",  prov_potreblenie)
    sys.exit()


wb.close()
# for v in pokk_2.values(): print(v[0],v[2])
